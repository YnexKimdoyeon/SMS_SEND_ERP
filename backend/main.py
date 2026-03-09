from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from datetime import date, datetime
from typing import Optional
import requests
import os
import io
from dotenv import load_dotenv

from database import engine, get_db, Base
from models import User, Payment, SmsLog, Setting

load_dotenv()

Base.metadata.create_all(bind=engine)

app = FastAPI(title="입금 알림 문자 발송 시스템")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://ynex3.mycafe24.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# === Schemas ===

class UserCreate(BaseModel):
    name: str
    phone: str

class UserResponse(BaseModel):
    id: int
    name: str
    phone: str
    created_at: datetime

    class Config:
        from_attributes = True

class PaymentCreate(BaseModel):
    user_id: int
    due_date: date
    amount: int
    memo: str = ""

class PaymentUpdate(BaseModel):
    is_paid: Optional[bool] = None
    due_date: Optional[date] = None
    amount: Optional[int] = None
    memo: Optional[str] = None

class PaymentResponse(BaseModel):
    id: int
    user_id: int
    due_date: date
    amount: int
    memo: str
    is_paid: bool
    sms_sent: bool
    created_at: datetime
    user: UserResponse

    class Config:
        from_attributes = True

class SMSSend(BaseModel):
    payment_id: int
    message: Optional[str] = None


# === Helper: SMS 발송 + 로그 기록 ===

def _send_sms_and_log(payment: Payment, msg: str, db: Session) -> bool:
    """문자 발송 후 로그 기록. 성공 시 True 반환."""
    api_key = os.getenv("ALIGO_API_KEY")
    user_id_env = os.getenv("ALIGO_USER_ID")
    sender = os.getenv("ALIGO_SENDER")

    if not all([api_key, user_id_env, sender]):
        db.add(SmsLog(
            payment_id=payment.id, user_name=payment.user.name,
            phone=payment.user.phone, message=msg,
            status="fail", error="알리고 API 설정 없음",
        ))
        db.commit()
        return False

    try:
        response = requests.post(
            "https://apis.aligo.in/send/",
            data={
                "key": api_key,
                "user_id": user_id_env,
                "sender": sender,
                "receiver": payment.user.phone,
                "msg": msg,
                "msg_type": "LMS" if len(msg.encode("utf-8")) > 90 else "SMS",
                "title": "입금 안내" if len(msg.encode("utf-8")) > 90 else "",
            },
        )
        result = response.json()

        if result.get("result_code") == "1":
            payment.sms_sent = True
            db.add(SmsLog(
                payment_id=payment.id, user_name=payment.user.name,
                phone=payment.user.phone, message=msg, status="success",
            ))
            db.commit()
            return True
        else:
            error_msg = result.get("message", "알 수 없는 오류")
            db.add(SmsLog(
                payment_id=payment.id, user_name=payment.user.name,
                phone=payment.user.phone, message=msg,
                status="fail", error=error_msg,
            ))
            db.commit()
            return False
    except requests.RequestException as e:
        db.add(SmsLog(
            payment_id=payment.id, user_name=payment.user.name,
            phone=payment.user.phone, message=msg,
            status="fail", error=str(e),
        ))
        db.commit()
        return False


# === User Endpoints ===

@app.get("/api/users", response_model=list[UserResponse])
def get_users(db: Session = Depends(get_db)):
    return db.query(User).order_by(User.name).all()

@app.post("/api/users", response_model=UserResponse)
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    db_user = User(name=user.name, phone=user.phone)
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.put("/api/users/{user_id}", response_model=UserResponse)
def update_user(user_id: int, user: UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    db_user.name = user.name
    db_user.phone = user.phone
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    db.delete(db_user)
    db.commit()
    return {"message": "삭제 완료"}


# === Payment Endpoints ===

@app.get("/api/payments", response_model=list[PaymentResponse])
def get_payments(year: Optional[int] = None, month: Optional[int] = None, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Payment)
    if user_id:
        query = query.filter(Payment.user_id == user_id)
    if year and month:
        start = date(year, month, 1)
        if month == 12:
            end = date(year + 1, 1, 1)
        else:
            end = date(year, month + 1, 1)
        query = query.filter(Payment.due_date >= start, Payment.due_date < end)
    return query.order_by(Payment.due_date).all()

@app.post("/api/payments", response_model=PaymentResponse)
def create_payment(payment: PaymentCreate, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == payment.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    db_payment = Payment(
        user_id=payment.user_id,
        due_date=payment.due_date,
        amount=payment.amount,
        memo=payment.memo,
    )
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    return db_payment

@app.put("/api/payments/{payment_id}", response_model=PaymentResponse)
def update_payment(payment_id: int, payment: PaymentUpdate, db: Session = Depends(get_db)):
    db_payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="입금 정보를 찾을 수 없습니다")
    if payment.is_paid is not None:
        db_payment.is_paid = payment.is_paid
    if payment.due_date is not None:
        db_payment.due_date = payment.due_date
    if payment.amount is not None:
        db_payment.amount = payment.amount
    if payment.memo is not None:
        db_payment.memo = payment.memo
    db.commit()
    db.refresh(db_payment)
    return db_payment

@app.delete("/api/payments/{payment_id}")
def delete_payment(payment_id: int, db: Session = Depends(get_db)):
    db_payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="입금 정보를 찾을 수 없습니다")
    db.delete(db_payment)
    db.commit()
    return {"message": "삭제 완료"}


# === SMS Endpoints ===

@app.post("/api/sms/send")
def send_sms(data: SMSSend, db: Session = Depends(get_db)):
    payment = db.query(Payment).filter(Payment.id == data.payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="입금 정보를 찾을 수 없습니다")

    if not data.message:
        msg = (
            f"[입금 안내] {payment.user.name}님, "
            f"{payment.due_date.strftime('%Y년 %m월 %d일')} "
            f"입금 예정일입니다. 금액: {payment.amount:,}원"
        )
        if payment.memo:
            msg += f" ({payment.memo})"
    else:
        msg = data.message

    success = _send_sms_and_log(payment, msg, db)
    if success:
        return {"message": "문자 발송 완료"}
    else:
        raise HTTPException(status_code=400, detail="문자 발송 실패")


@app.post("/api/sms/send-bulk")
def send_bulk_sms(db: Session = Depends(get_db)):
    payments = (
        db.query(Payment)
        .filter(Payment.is_paid == False, Payment.sms_sent == False)
        .all()
    )
    if not payments:
        return {"message": "발송할 대상이 없습니다", "sent": 0}

    sent_count = 0
    for payment in payments:
        msg = (
            f"[입금 안내] {payment.user.name}님, "
            f"{payment.due_date.strftime('%Y년 %m월 %d일')} "
            f"입금 예정일입니다. 금액: {payment.amount:,}원"
        )
        if payment.memo:
            msg += f" ({payment.memo})"
        if _send_sms_and_log(payment, msg, db):
            sent_count += 1

    return {"message": f"{sent_count}건 발송 완료", "sent": sent_count}


# === SMS 발송 이력 ===

@app.get("/api/sms/logs")
def get_sms_logs(page: int = 1, size: int = 50, db: Session = Depends(get_db)):
    total = db.query(func.count(SmsLog.id)).scalar()
    logs = (
        db.query(SmsLog)
        .order_by(SmsLog.sent_at.desc())
        .offset((page - 1) * size)
        .limit(size)
        .all()
    )
    return {
        "total": total,
        "page": page,
        "size": size,
        "logs": [
            {
                "id": log.id,
                "user_name": log.user_name,
                "phone": log.phone,
                "message": log.message,
                "status": log.status,
                "error": log.error,
                "sent_at": log.sent_at.isoformat() if log.sent_at else None,
            }
            for log in logs
        ],
    }


# === 대시보드 ===

@app.get("/api/dashboard")
def get_dashboard(year: Optional[int] = None, month: Optional[int] = None, db: Session = Depends(get_db)):
    if not year or not month:
        today = date.today()
        year, month = today.year, today.month

    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    payments = db.query(Payment).filter(
        Payment.due_date >= start, Payment.due_date < end
    ).all()

    total_count = len(payments)
    total_amount = sum(p.amount for p in payments)
    paid_count = sum(1 for p in payments if p.is_paid)
    paid_amount = sum(p.amount for p in payments if p.is_paid)
    unpaid_count = total_count - paid_count
    unpaid_amount = total_amount - paid_amount
    sms_sent_count = sum(1 for p in payments if p.sms_sent)

    # 오늘 받아야 할 금액
    today = date.today()
    today_payments = [p for p in payments if p.due_date == today and not p.is_paid]
    today_count = len(today_payments)
    today_amount = sum(p.amount for p in today_payments)

    user_count = db.query(func.count(User.id)).scalar()

    # 최근 문자 발송 5건
    recent_logs = (
        db.query(SmsLog)
        .order_by(SmsLog.sent_at.desc())
        .limit(5)
        .all()
    )

    return {
        "year": year,
        "month": month,
        "total_count": total_count,
        "total_amount": total_amount,
        "paid_count": paid_count,
        "paid_amount": paid_amount,
        "unpaid_count": unpaid_count,
        "unpaid_amount": unpaid_amount,
        "sms_sent_count": sms_sent_count,
        "today_count": today_count,
        "today_amount": today_amount,
        "user_count": user_count,
        "recent_logs": [
            {
                "user_name": log.user_name,
                "status": log.status,
                "sent_at": log.sent_at.isoformat() if log.sent_at else None,
            }
            for log in recent_logs
        ],
    }


# === 엑셀 다운로드 ===

@app.get("/api/export/excel")
def export_excel(year: Optional[int] = None, month: Optional[int] = None, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not year or not month:
        today = date.today()
        year, month = today.year, today.month

    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    payments = (
        db.query(Payment)
        .filter(Payment.due_date >= start, Payment.due_date < end)
        .order_by(Payment.due_date)
        .all()
    )

    wb = Workbook()

    # --- 입금 내역 시트 ---
    ws = wb.active
    ws.title = f"{year}년 {month}월 입금내역"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    paid_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    unpaid_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["번호", "이름", "전화번호", "입금일", "금액", "메모", "입금여부", "문자발송"]
    col_widths = [6, 12, 16, 14, 14, 20, 10, 10]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + i)].width = w

    for idx, p in enumerate(payments, 1):
        row = idx + 1
        values = [
            idx,
            p.user.name,
            p.user.phone,
            p.due_date.strftime("%Y-%m-%d"),
            p.amount,
            p.memo,
            "입금완료" if p.is_paid else "미입금",
            "발송됨" if p.sms_sent else "미발송",
        ]
        fill = paid_fill if p.is_paid else unpaid_fill
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.fill = fill
            if col == 5:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col in (1, 7, 8):
                cell.alignment = Alignment(horizontal="center")

    # 합계 행
    total_row = len(payments) + 2
    ws.cell(row=total_row, column=4, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(p.amount for p in payments)).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = "#,##0"

    paid_total_row = total_row + 1
    ws.cell(row=paid_total_row, column=4, value="입금완료").font = Font(bold=True, color="2E7D32")
    ws.cell(row=paid_total_row, column=5, value=sum(p.amount for p in payments if p.is_paid)).font = Font(bold=True, color="2E7D32")
    ws.cell(row=paid_total_row, column=5).number_format = "#,##0"

    unpaid_total_row = total_row + 2
    ws.cell(row=unpaid_total_row, column=4, value="미입금").font = Font(bold=True, color="C62828")
    ws.cell(row=unpaid_total_row, column=5, value=sum(p.amount for p in payments if not p.is_paid)).font = Font(bold=True, color="C62828")
    ws.cell(row=unpaid_total_row, column=5).number_format = "#,##0"

    # --- 문자 발송 이력 시트 ---
    ws2 = wb.create_sheet(title="문자 발송 이력")
    sms_headers = ["번호", "이름", "전화번호", "발송내용", "상태", "발송시간", "오류"]
    sms_widths = [6, 12, 16, 40, 10, 20, 30]

    for i, (h, w) in enumerate(zip(sms_headers, sms_widths), 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws2.column_dimensions[chr(64 + i)].width = w

    logs = db.query(SmsLog).order_by(SmsLog.sent_at.desc()).limit(500).all()
    for idx, log in enumerate(logs, 1):
        row = idx + 1
        values = [
            idx,
            log.user_name,
            log.phone,
            log.message,
            "성공" if log.status == "success" else "실패",
            log.sent_at.strftime("%Y-%m-%d %H:%M") if log.sent_at else "",
            log.error or "",
        ]
        success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        fill = success_fill if log.status == "success" else fail_fill
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (1, 5):
                cell.alignment = Alignment(horizontal="center")
            if col == 5:
                cell.fill = fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"payment_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# === 설정 API ===

@app.get("/api/settings/send-time")
def get_send_time(db: Session = Depends(get_db)):
    setting = db.query(Setting).filter(Setting.key == "send_time").first()
    return {"send_time": setting.value if setting else "09:00"}

@app.put("/api/settings/send-time")
def update_send_time(data: dict, db: Session = Depends(get_db)):
    time_str = data.get("send_time", "09:00")
    try:
        h, m = map(int, time_str.split(":"))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            raise ValueError
    except ValueError:
        raise HTTPException(status_code=400, detail="올바른 시간 형식이 아닙니다 (HH:MM)")

    setting = db.query(Setting).filter(Setting.key == "send_time").first()
    if setting:
        setting.value = time_str
    else:
        db.add(Setting(key="send_time", value=time_str))
    db.commit()

    reschedule_job(h, m)
    return {"message": f"발송 시간이 {time_str}으로 설정되었습니다", "send_time": time_str}


# === 스케줄러 ===

from apscheduler.schedulers.background import BackgroundScheduler

def auto_send_daily():
    db = next(get_db())
    try:
        today = date.today()
        payments = (
            db.query(Payment)
            .filter(Payment.due_date == today, Payment.is_paid == False, Payment.sms_sent == False)
            .all()
        )
        if not payments:
            return

        for payment in payments:
            msg = (
                f"[입금 안내] {payment.user.name}님, "
                f"오늘({today.strftime('%Y.%m.%d')}) 입금 예정일입니다. "
                f"금액: {payment.amount:,}원"
            )
            if payment.memo:
                msg += f" ({payment.memo})"
            _send_sms_and_log(payment, msg, db)
    finally:
        db.close()

scheduler = BackgroundScheduler()

def reschedule_job(hour: int, minute: int):
    if scheduler.get_job("daily_sms"):
        scheduler.remove_job("daily_sms")
    scheduler.add_job(auto_send_daily, "cron", hour=hour, minute=minute, id="daily_sms")

def init_scheduler():
    db = next(get_db())
    try:
        setting = db.query(Setting).filter(Setting.key == "send_time").first()
        time_str = setting.value if setting else "09:00"
        h, m = map(int, time_str.split(":"))
        reschedule_job(h, m)
    finally:
        db.close()
    scheduler.start()

init_scheduler()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
